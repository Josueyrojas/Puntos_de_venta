# metodos.py - Logica del negocio y funciones del punto de venta

from datetime import datetime
from typing import List, Dict, Optional, Tuple
import json
import os
import csv
import re
from openpyxl import Workbook, load_workbook

def validar_email(email: str) -> bool:
    """Valida el formato de un correo electronico"""
    patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(patron, email) is not None


def validar_rfc(rfc: str) -> bool:
    """Valida el formato de un RFC mexicano"""
    patron = r'^[A-Z&]{3,4}[0-9]{6}[A-Z0-9]{3}$'
    return re.match(patron, rfc.upper()) is not None


def validar_numero(valor, tipo: str = "float") -> bool:
    """Valida si un valor es un numero del tipo especificado"""
    try:
        if tipo == "int":
            int(valor)
        else:  # float
            float(valor)
        return True
    except (ValueError, TypeError):
        return False


def formatear_moneda(valor: float) -> str:
    """Formatea un valor como moneda"""
    return f"${valor:,.2f}"


def obtener_regimenes_fiscales() -> List[Dict[str, str]]:
    """Retorna la lista de regimenes fiscales disponibles en Mexico"""
    return [
        {'clave': '601', 'descripcion': 'General de Ley Personas Morales'},
        {'clave': '603', 'descripcion': 'Personas Morales con Fines no Lucrativos'},
        {'clave': '605', 'descripcion': 'Sueldos y Salarios e Ingresos Asimilados'},
        {'clave': '606', 'descripcion': 'Actividades Empresariales con Ingresos por Arrendamiento'},
        {'clave': '607', 'descripcion': 'Ingresos por Dividendos (socios y accionistas)'},
        {'clave': '608', 'descripcion': 'Ingresos Derivados de Construcciones'},
        {'clave': '609', 'descripcion': 'Otras Fuentes de Ingreso'},
        {'clave': '610', 'descripcion': 'Residentes en el Extranjero sin Establecimiento Permanente'},
        {'clave': '614', 'descripcion': 'Ingresos por Intereses'},
        {'clave': '616', 'descripcion': 'Sin Obligacion Fiscal'},
        {'clave': '620', 'descripcion': 'Sociedades Cooperativas'},
        {'clave': '621', 'descripcion': 'Agrupaciones en Participacion'},
        {'clave': '622', 'descripcion': 'Sociedades y Asociaciones Civiles'},
        {'clave': '623', 'descripcion': 'Asociacion en Participacion'},
        {'clave': '624', 'descripcion': 'Fideicomiso'},
        {'clave': '625', 'descripcion': 'Asociaciones Ganaderas'},
        {'clave': '626', 'descripcion': 'Sociedades y Asociaciones Civiles dedicadas a actividades agricolas'},
        {'clave': '627', 'descripcion': 'Sociedades Mercantiles y Asociaciones Civiles'},
        {'clave': '628', 'descripcion': 'Personas Fisicas con Actividades Empresariales'},
        {'clave': '629', 'descripcion': 'Personas Fisicas con Actividades Profesionales'},
        {'clave': '630', 'descripcion': 'Ingresos por Intereses, Dividendos, Arrendamientos, Regalías y otros Ingresos del Extranjero'},
    ]


def obtener_usos_cfdi() -> List[Dict[str, str]]:
    """Retorna la lista de usos de CFDI disponibles"""
    return [
        {'clave': 'G01', 'descripcion': 'Adquisición de mercancias'},
        {'clave': 'G02', 'descripcion': 'Devoluciones, descuentos o bonificaciones'},
        {'clave': 'G03', 'descripcion': 'Gastos en general'},
        {'clave': 'I01', 'descripcion': 'Construcciones'},
        {'clave': 'I02', 'descripcion': 'Mobilario y equipo de oficina por inversiones'},
        {'clave': 'I03', 'descripcion': 'Equipo de transporte'},
        {'clave': 'I04', 'descripcion': 'Equipo de computo y accesorios'},
        {'clave': 'I05', 'descripcion': 'Dados, troqueles, moldes, matrices y herramental'},
        {'clave': 'I06', 'descripcion': 'Depósitos en garantia y pólizas de seguros'},
        {'clave': 'I07', 'descripcion': 'Adecuacion del inmueble'},
        {'clave': 'I08', 'descripcion': 'Equipamiento de establecimientos'},
        {'clave': 'D01', 'descripcion': 'Honorarios médicos'},
        {'clave': 'D02', 'descripcion': 'Honorarios dentales'},
        {'clave': 'D03', 'descripcion': 'Honorarios hospitalarios'},
        {'clave': 'D04', 'descripcion': 'Gastos médicos por incapacidad o discapacidad'},
        {'clave': 'D05', 'descripcion': 'Transporte escolar obligatorio'},
        {'clave': 'D06', 'descripcion': 'Retencion agraria'},
        {'clave': 'D07', 'descripcion': 'Pago por servicios educativos'},
        {'clave': 'D08', 'descripcion': 'Ritmo cardiaco implantable'},
        {'clave': 'D09', 'descripcion': 'Feretro'},
        {'clave': 'D10', 'descripcion': 'Donativo'},
        {'clave': 'D11', 'descripcion': 'Pago del seguro de sobrantes en gasolineras'},
        {'clave': 'D12', 'descripcion': 'Pagos referentes a subsistemas de salud'},
        {'clave': 'D13', 'descripcion': 'Pagos por nutrición'},
        {'clave': 'D14', 'descripcion': 'Pago referente a vivienda'},
        {'clave': 'D15', 'descripcion': 'Pago referente a servicios funerarios'},
    ]

class ProductoVenta:
    """Clase para representar un producto en la venta"""
    def __init__(self, codigo_barras: str, datos_producto: Dict, cantidad: int = 1):
        self.codigo_barras = codigo_barras
        self.codigo = datos_producto.get('codigo', '')
        self.numero_producto = datos_producto.get('numero_producto', '')
        self.nombre = datos_producto.get('nombre', '')
        self.descripcion = datos_producto.get('descripcion', '')
        self.clasificacion = datos_producto.get('clasificacion', '')
        self.precio_minorista = float(datos_producto.get('precio_minorista', 0))
        self.precio_mayoreo = float(datos_producto.get('precio_mayoreo', 0))
        self.costo = float(datos_producto.get('costo', 0))
        self.proveedor = datos_producto.get('proveedor', '')
        self.unidad = datos_producto.get('unidad', 'pz')
        self.fabricante = datos_producto.get('fabricante', '')
        self.tipo = datos_producto.get('tipo', '')
        self.codigoA = datos_producto.get('codigoA', '')
        self.codigoB = datos_producto.get('codigoB', '')
        self.codigoC = datos_producto.get('codigoC', '')
        self.cantidad = cantidad
    
    def obtener_precio_unitario(self) -> float:
        """Obtiene el precio unitario según la cantidad (regla: 6+ piezas = mayoreo)"""
        if self.cantidad >= 6:
            return self.precio_mayoreo
        return self.precio_minorista
    
    def subtotal(self) -> float:
        """Calcula el subtotal del producto aplicando precio según cantidad"""
        precio_unitario = self.obtener_precio_unitario()
        return precio_unitario * self.cantidad
    
    def to_dict(self) -> Dict:
        """Convierte el producto a diccionario"""
        precio_unitario = self.obtener_precio_unitario()
        return {
            "codigo_barras": self.codigo_barras,
            "codigo": self.codigo,
            "numero_producto": self.numero_producto,
            "nombre": self.nombre,
            "descripcion": self.descripcion,
            "clasificacion": self.clasificacion,
            "precio_minorista": self.precio_minorista,
            "precio_mayoreo": self.precio_mayoreo,
            "precio_aplicado": precio_unitario,
            "costo": self.costo,
            "proveedor": self.proveedor,
            "unidad": self.unidad,
            "fabricante": self.fabricante,
            "tipo": self.tipo,
            "codigoA": self.codigoA,
            "codigoB": self.codigoB,
            "codigoC": self.codigoC,
            "cantidad": self.cantidad,
            "subtotal": self.subtotal(),
            "es_mayoreo": self.cantidad >= 6
        }


class Gestor_Inventario:
    """Clase para gestionar el inventario de productos"""
    def __init__(self):
        self.productos = {}
        self.cargar_inventario()
    
    def cargar_inventario(self, archivo: str = 'inventario.json') -> None:
        """Carga el inventario desde archivo JSON"""
        if os.path.exists(archivo):
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    self.productos = json.load(f)
            except Exception as e:
                print(f"Error al cargar inventario: {e}")
    
    def guardar_inventario(self, archivo: str = 'inventario.json') -> None:
        """Guarda el inventario en archivo JSON"""
        try:
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump(self.productos, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error al guardar inventario: {e}")
    
    def obtener_producto(self, codigo: str) -> Optional[Dict]:
        """Obtiene un producto del inventario"""
        return self.productos.get(codigo)
    
    def obtener_todos(self) -> Dict:
        """Obtiene todos los productos"""
        return self.productos
    
    def actualizar_stock(self, codigo: str, cantidad: int) -> bool:
        """Actualiza el stock de un producto"""
        if codigo in self.productos:
            self.productos[codigo]['stock'] = self.productos[codigo].get('stock', 0) + cantidad
            self.guardar_inventario()
            return True
        return False
    
    def agregar_producto(self, 
                        codigo_barras: str,
                        codigo: str,
                        numero_producto: str,
                        nombre: str,
                        descripcion: str = "",
                        clasificacion: str = "",
                        precio_minorista: float = 0,
                        precio_mayoreo: float = 0,
                        costo: float = 0,
                        proveedor: str = "",
                        unidad: str = "pz",
                        fabricante: str = "",
                        tipo: str = "",
                        codigoA: str = "",
                        codigoB: str = "",
                        codigoC: str = "",
                        stock: int = 0) -> None:
        """Agrega un producto al inventario con todos los nuevos campos"""
        self.productos[codigo_barras] = {
            'codigo': codigo,
            'numero_producto': numero_producto,
            'nombre': nombre,
            'descripcion': descripcion,
            'clasificacion': clasificacion,
            'precio_minorista': precio_minorista,
            'precio_mayoreo': precio_mayoreo,
            'costo': costo,
            'proveedor': proveedor,
            'unidad': unidad,
            'fabricante': fabricante,
            'tipo': tipo,
            'codigoA': codigoA,
            'codigoB': codigoB,
            'codigoC': codigoC,
            'stock': stock,
            'fecha_creacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        self.guardar_inventario()
    
    def editar_producto(self,
                       codigo_barras: str,
                       codigo: str,
                       numero_producto: str,
                       nombre: str,
                       descripcion: str = "",
                       clasificacion: str = "",
                       precio_minorista: float = 0,
                       precio_mayoreo: float = 0,
                       costo: float = 0,
                       proveedor: str = "",
                       unidad: str = "pz",
                       fabricante: str = "",
                       tipo: str = "",
                       codigoA: str = "",
                       codigoB: str = "",
                       codigoC: str = "",
                       stock: int = 0) -> None:
        """Edita un producto del inventario"""
        if codigo_barras in self.productos:
            self.productos[codigo_barras] = {
                'codigo': codigo,
                'numero_producto': numero_producto,
                'nombre': nombre,
                'descripcion': descripcion,
                'clasificacion': clasificacion,
                'precio_minorista': precio_minorista,
                'precio_mayoreo': precio_mayoreo,
                'costo': costo,
                'proveedor': proveedor,
                'unidad': unidad,
                'fabricante': fabricante,
                'tipo': tipo,
                'codigoA': codigoA,
                'codigoB': codigoB,
                'codigoC': codigoC,
                'stock': stock,
                'fecha_creacion': self.productos[codigo_barras].get('fecha_creacion', 
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.guardar_inventario()
    
    def eliminar_producto(self, codigo_barras: str) -> bool:
        """Elimina un producto del inventario"""
        if codigo_barras in self.productos:
            del self.productos[codigo_barras]
            self.guardar_inventario()
            return True
        return False
    
    # MODIFICADO: Ahora busca principalmente por descripción
    def buscar_producto_por_descripcion(self, descripcion: str) -> List[Tuple[str, Dict]]:
        """Busca productos por descripción (ahora el método principal para punto de venta)"""
        resultados = []
        descripcion_lower = descripcion.lower()
        
        for codigo_barras, producto in self.productos.items():
            # Buscar en múltiples campos relacionados con la descripción
            if (descripcion_lower in producto.get('descripcion', '').lower() or
                descripcion_lower in producto.get('nombre', '').lower() or
                descripcion_lower in producto.get('clasificacion', '').lower() or
                descripcion_lower in producto.get('fabricante', '').lower() or
                descripcion_lower in producto.get('tipo', '').lower()):
                
                resultados.append((codigo_barras, producto))
        
        return resultados
    
    # Método auxiliar para búsqueda por código (para otras funcionalidades)
    def buscar_producto_por_codigo(self, codigo_busqueda: str) -> Optional[Tuple[str, Dict]]:
        """Busca un producto por cualquier código: barras, producto, A, B o C"""
        # 1. Buscar por código de barras (clave principal)
        if codigo_busqueda in self.productos:
            return (codigo_busqueda, self.productos[codigo_busqueda])
        
        # 2. Buscar en los demás campos de código
        for cb, producto in self.productos.items():
            if (producto.get('codigo') == codigo_busqueda or
                producto.get('numero_producto') == codigo_busqueda or
                producto.get('codigoA') == codigo_busqueda or
                producto.get('codigoB') == codigo_busqueda or
                producto.get('codigoC') == codigo_busqueda):
                return (cb, producto)
        
        return None
    
    def tiene_stock(self, codigo_barras: str, cantidad: int) -> bool:
        """Verifica si hay stock disponible"""
        if codigo_barras in self.productos:
            return self.productos[codigo_barras].get('stock', 0) >= cantidad
        return False
    
    def buscar_productos_por_nombre(self, termino: str) -> List[Tuple[str, Dict]]:
        """Busca productos por nombre (método legacy)"""
        resultados = []
        termino_lower = termino.lower()
        for codigo_barras, producto in self.productos.items():
            if termino_lower in producto.get('nombre', '').lower():
                resultados.append((codigo_barras, producto))
        return resultados
    
    def importar_inventario(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa inventario desde un archivo JSON"""
        try:
            with open(ruta, 'r', encoding='utf-8') as f:
                datos_importados = json.load(f)
            
            if sobrescribir:
                self.productos = datos_importados
            else:
                self.productos.update(datos_importados)
            
            self.guardar_inventario()
            return (True, f"Inventario importado: {len(datos_importados)} productos")
        except Exception as e:
            return (False, f"Error al importar: {str(e)}")
    
    def exportar_inventario(self, ruta: str) -> bool:
        """Exporta el inventario a un archivo JSON"""
        try:
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(self.productos, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Error al exportar: {e}")
            return False
    
    def obtener_producto_para_venta(self, codigo_barras: str, cantidad: int = 1) -> Optional[ProductoVenta]:
        """Obtiene un producto listo para agregar a venta"""
        if codigo_barras in self.productos:
            producto_data = self.productos[codigo_barras].copy()
            # Verificar stock disponible
            if producto_data.get('stock', 0) >= cantidad:
                return ProductoVenta(codigo_barras, producto_data, cantidad)
        return None
    
    def buscar_productos_avanzado(self, criterio: str) -> List[Tuple[str, Dict]]:
        """Busca productos por múltiples criterios (descripción, nombre, código, etc.)"""
        resultados = []
        criterio_lower = criterio.lower()
        
        for codigo_barras, producto in self.productos.items():
            # Buscar en todos los campos relevantes
            if (criterio_lower in producto.get('descripcion', '').lower() or
                criterio_lower in producto.get('nombre', '').lower() or
                criterio_lower in producto.get('clasificacion', '').lower() or
                criterio_lower in producto.get('fabricante', '').lower() or
                criterio_lower in str(producto.get('codigo', '')).lower() or
                criterio_lower in str(producto.get('numero_producto', '')).lower() or
                criterio_lower in str(producto.get('codigoA', '')).lower() or
                criterio_lower in str(producto.get('codigoB', '')).lower() or
                criterio_lower in str(producto.get('codigoC', '')).lower()):
                
                resultados.append((codigo_barras, producto))
        
        return resultados
    
    def generar_plantilla_csv(self, ruta: str = 'plantilla_inventario.csv') -> Tuple[bool, str]:
        """Genera un archivo CSV de plantilla con los campos requeridos para inventario"""
        try:
            campos = [
                'Codigo Barras',
                'Codigo',
                'Numero Producto',
                'Nombre',
                'Descripcion',
                'Clasificacion',
                'Precio Minorista',
                'Precio Mayoreo',
                'Costo',
                'Proveedor',
                'Unidad',
                'Fabricante',
                'Tipo',
                'Codigo A',
                'Codigo B',
                'Codigo C',
                'Stock'
            ]
            
            with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(campos)
                # Agregar una fila de ejemplo
                writer.writerow([
                    '7501234567890',
                    'PROD001',
                    '001',
                    'Producto Ejemplo',
                    'Descripción del producto ejemplo',
                    'Herramientas',
                    '100.00',
                    '85.00',
                    '50.00',
                    'Distribuidor Central',
                    'pz',
                    'Fabricante Ejemplo',
                    'Tipo A',
                    'COD-A-001',
                    'COD-B-001',
                    'COD-C-001',
                    '50'
                ])
            
            return (True, f"Plantilla generada: {ruta}")
        
        except Exception as e:
            return (False, f"Error al generar plantilla: {str(e)}")


class GestorVentas:
    """Clase para gestionar las ventas con nueva lógica de precios"""
    def __init__(self, gestor_inventario: Gestor_Inventario = None):
        self.productos_venta = []
        self.historial_ventas = []
        self.gestor_inventario = gestor_inventario
        self.cargar_historial()
        self.numero_folio = len(self.historial_ventas) + 1
    
    def cargar_historial(self, archivo: str = 'ventas.json') -> None:
        """Carga el historial de ventas desde archivo JSON"""
        if os.path.exists(archivo):
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    self.historial_ventas = json.load(f)
            except Exception as e:
                print(f"Error al cargar ventas: {e}")
    
    def guardar_historial(self, archivo: str = 'ventas.json') -> None:
        """Guarda el historial de ventas en archivo JSON"""
        try:
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump(self.historial_ventas, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error al guardar ventas: {e}")
    
    # MODIFICADO: Nuevo método para agregar producto por descripción
    def agregar_producto_por_descripcion(self, descripcion: str, cantidad: int = 1) -> Tuple[bool, str, Optional[ProductoVenta]]:
        """Agrega un producto a la venta buscando por descripción"""
        if not self.gestor_inventario:
            return (False, "No hay gestor de inventario configurado", None)
        
        if not descripcion.strip():
            return (False, "La descripción no puede estar vacía", None)
        
        if cantidad <= 0:
            return (False, "La cantidad debe ser mayor a 0", None)
        
        # Buscar productos por descripción
        resultados = self.gestor_inventario.buscar_producto_por_descripcion(descripcion)
        
        if not resultados:
            return (False, f"No se encontraron productos con: '{descripcion}'", None)
        
        # Si hay múltiples resultados, tomamos el primero
        # (En una interfaz gráfica, se mostraría una lista para seleccionar)
        codigo_barras, producto_data = resultados[0]
        
        # Verificar stock
        if not self.gestor_inventario.tiene_stock(codigo_barras, cantidad):
            stock_disponible = producto_data.get('stock', 0)
            return (False, f"Stock insuficiente. Disponible: {stock_disponible}", None)
        
        # Crear producto para venta
        producto_venta = ProductoVenta(codigo_barras, producto_data, cantidad)
        
        # Verificar si el producto ya existe en la venta
        for prod in self.productos_venta:
            if prod.codigo_barras == codigo_barras:
                # Calcular nueva cantidad total
                nueva_cantidad = prod.cantidad + cantidad
                
                # Verificar stock con la nueva cantidad
                if not self.gestor_inventario.tiene_stock(codigo_barras, nueva_cantidad):
                    stock_disponible = producto_data.get('stock', 0)
                    return (False, f"Stock insuficiente para la cantidad total. Disponible: {stock_disponible}", None)
                
                prod.cantidad = nueva_cantidad
                mensaje = f"Producto actualizado: {prod.nombre} - {nueva_cantidad} {prod.unidad}"
                
                # Calcular si ahora aplica precio mayoreo
                if nueva_cantidad >= 6 and prod.cantidad < 6:
                    mensaje += f" - Ahora aplica precio mayoreo"
                
                return (True, mensaje, producto_venta)
        
        # Si no existe, agregarlo
        self.productos_venta.append(producto_venta)
        mensaje = f"Producto agregado: {producto_venta.nombre} - {cantidad} {producto_venta.unidad}"
        
        # Indicar si aplica precio mayoreo
        if cantidad >= 6:
            mensaje += f" - Aplica precio mayoreo"
        
        return (True, mensaje, producto_venta)
    
    # Método original para agregar producto por código (para compatibilidad)
    def agregar_producto(self, producto: ProductoVenta) -> Tuple[bool, str]:
        """Agrega un producto a la venta actual, verificando stock"""
        if not self.gestor_inventario:
            return (False, "No hay gestor de inventario configurado")
        
        # Verificar stock disponible
        if not self.gestor_inventario.tiene_stock(producto.codigo_barras, producto.cantidad):
            producto_data = self.gestor_inventario.obtener_producto(producto.codigo_barras)
            stock_disponible = producto_data.get('stock', 0) if producto_data else 0
            return (False, f"Stock insuficiente. Disponible: {stock_disponible}")
        
        # Verificar si el producto ya existe en la venta
        for prod in self.productos_venta:
            if prod.codigo_barras == producto.codigo_barras:
                # Calcular nueva cantidad total
                nueva_cantidad = prod.cantidad + producto.cantidad
                
                # Verificar stock con la nueva cantidad
                if not self.gestor_inventario.tiene_stock(producto.codigo_barras, nueva_cantidad):
                    stock_disponible = producto_data.get('stock', 0) if producto_data else 0
                    return (False, f"Stock insuficiente para la cantidad total. Disponible: {stock_disponible}")
                
                prod.cantidad = nueva_cantidad
                return (True, f"Producto actualizado: {prod.nombre} - {nueva_cantidad} {prod.unidad}")
        
        self.productos_venta.append(producto)
        return (True, f"Producto agregado: {producto.nombre}")
    
    def eliminar_producto(self, codigo_barras: str) -> bool:
        """Elimina un producto de la venta actual"""
        for i, prod in enumerate(self.productos_venta):
            if prod.codigo_barras == codigo_barras:
                self.productos_venta.pop(i)
                return True
        return False
    
    def eliminar_producto_por_descripcion(self, descripcion: str) -> Tuple[bool, str]:
        """Elimina un producto de la venta buscando por descripción"""
        descripcion_lower = descripcion.lower()
        
        for i, prod in enumerate(self.productos_venta):
            if (descripcion_lower in prod.descripcion.lower() or
                descripcion_lower in prod.nombre.lower()):
                
                nombre_eliminado = prod.nombre
                self.productos_venta.pop(i)
                return (True, f"Producto eliminado: {nombre_eliminado}")
        
        return (False, f"No se encontró producto con: '{descripcion}'")
    
    def obtener_productos_venta(self) -> List[ProductoVenta]:
        """Obtiene los productos de la venta actual"""
        return self.productos_venta
    
    def calcular_subtotal(self) -> float:
        """Calcula el subtotal de la venta"""
        return sum(prod.subtotal() for prod in self.productos_venta)
    
    def calcular_iva(self, subtotal: float = None) -> float:
        """Calcula el IVA de la venta (16%)"""
        if subtotal is None:
            subtotal = self.calcular_subtotal()
        return subtotal * 0.16
    
    def calcular_total(self) -> float:
        """Calcula el total de la venta"""
        subtotal = self.calcular_subtotal()
        iva = self.calcular_iva(subtotal)
        return subtotal + iva
    
    def calcular_cambio(self, pago: float, total: float = None) -> float:
        """Calcula el cambio"""
        if total is None:
            total = self.calcular_total()
        return max(0, pago - total)
    
    def obtener_detalle_precios(self) -> Dict:
        """Obtiene detalle de precios aplicados en la venta"""
        detalle = {
            'productos_minorista': [],
            'productos_mayoreo': [],
            'total_minorista': 0,
            'total_mayoreo': 0,
            'descuento_total': 0
        }
        
        for prod in self.productos_venta:
            precio_unitario = prod.obtener_precio_unitario()
            subtotal = prod.subtotal()
            
            if prod.cantidad >= 6:
                detalle['productos_mayoreo'].append({
                    'nombre': prod.nombre,
                    'cantidad': prod.cantidad,
                    'precio_mayoreo': prod.precio_mayoreo,
                    'precio_minorista': prod.precio_minorista,
                    'subtotal': subtotal,
                    'ahorro': (prod.precio_minorista - prod.precio_mayoreo) * prod.cantidad
                })
                detalle['total_mayoreo'] += subtotal
                detalle['descuento_total'] += (prod.precio_minorista - prod.precio_mayoreo) * prod.cantidad
            else:
                detalle['productos_minorista'].append({
                    'nombre': prod.nombre,
                    'cantidad': prod.cantidad,
                    'precio': precio_unitario,
                    'subtotal': subtotal
                })
                detalle['total_minorista'] += subtotal
        
        return detalle
    
    def procesar_venta(self, pago: float, aplicar_descuento: bool = True) -> Tuple[bool, str, Optional[Dict]]:
        """Procesa la venta y la guarda en el historial"""
        if not self.productos_venta:
            return (False, "No hay productos en la venta", None)
        
        # Verificar stock antes de procesar la venta
        if self.gestor_inventario:
            for prod in self.productos_venta:
                if not self.gestor_inventario.tiene_stock(prod.codigo_barras, prod.cantidad):
                    return (False, f"Stock insuficiente para: {prod.nombre}", None)
        
        subtotal = self.calcular_subtotal()
        iva = self.calcular_iva(subtotal)
        total = subtotal + iva
        
        if pago < total:
            return (False, f"Pago insuficiente. Total: ${total:.2f}", None)
        
        cambio = self.calcular_cambio(pago, total)
        
        # Obtener detalle de precios aplicados
        detalle_precios = self.obtener_detalle_precios()
        
        venta = {
            'folio': self.numero_folio,
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'productos': [prod.to_dict() for prod in self.productos_venta],
            'subtotal': subtotal,
            'iva': iva,
            'total': total,
            'pago': pago,
            'cambio': cambio,
            'detalle_precios': detalle_precios,
            'productos_con_descuento': len(detalle_precios['productos_mayoreo']),
            'descuento_total': detalle_precios['descuento_total']
        }
        
        # Actualizar stock en inventario
        if self.gestor_inventario:
            for prod in self.productos_venta:
                self.gestor_inventario.actualizar_stock(prod.codigo_barras, -prod.cantidad)
        
        self.historial_ventas.append(venta)
        self.numero_folio += 1
        self.guardar_historial()
        
        # Limpiar venta actual después de procesar
        self.limpiar_venta()
        
        return (True, "Venta procesada exitosamente", venta)
    
    def limpiar_venta(self) -> None:
        """Limpia la venta actual"""
        self.productos_venta = []
    
    def obtener_historial(self) -> List[Dict]:
        """Obtiene el historial de ventas"""
        return self.historial_ventas
    
    def buscar_venta_por_folio(self, folio: int) -> Optional[Dict]:
        """Busca una venta por folio"""
        for venta in self.historial_ventas:
            if venta.get('folio') == folio:
                return venta
        return None
    
    def buscar_ventas_por_descripcion(self, descripcion: str) -> List[Dict]:
        """Busca ventas por descripción de productos"""
        resultados = []
        descripcion_lower = descripcion.lower()
        
        for venta in self.historial_ventas:
            for producto in venta.get('productos', []):
                if (descripcion_lower in producto.get('descripcion', '').lower() or
                    descripcion_lower in producto.get('nombre', '').lower()):
                    resultados.append(venta)
                    break  # Si encontramos en un producto, agregamos la venta y salimos
        
        return resultados
    
    def registrar_venta(self, venta: Dict) -> None:
        """Registra una venta (metodo legacy)"""
        self.historial_ventas.append(venta)
    
    def obtener_ventas(self) -> List[Dict]:
        """Obtiene el listado de ventas"""
        return self.historial_ventas
    
    def set_gestor_inventario(self, gestor_inventario: Gestor_Inventario) -> None:
        """Establece el gestor de inventario para validaciones de stock"""
        self.gestor_inventario = gestor_inventario


class GestorProveedores:
    """Clase para gestionar los proveedores con todos los campos actualizados"""
    def __init__(self):
        self.proveedores = {}
        self.cargar_proveedores()
    
    def cargar_proveedores(self, archivo: str = 'proveedores.json') -> None:
        """Carga los proveedores desde archivo JSON"""
        if os.path.exists(archivo):
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    self.proveedores = json.load(f)
            except Exception as e:
                print(f"Error al cargar proveedores: {e}")
    
    def guardar_proveedores(self, archivo: str = 'proveedores.json') -> None:
        """Guarda los proveedores en archivo JSON"""
        try:
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump(self.proveedores, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error al guardar proveedores: {e}")
    
    def obtener_proveedor(self, id_proveedor: str) -> Optional[Dict]:
        """Obtiene un proveedor por ID"""
        return self.proveedores.get(id_proveedor.upper())
    
    def obtener_todos(self) -> Dict:
        """Obtiene todos los proveedores"""
        return self.proveedores
    
    def obtener_lista_simplificada(self) -> List[Dict]:
        """Obtiene lista de proveedores para combos/selects"""
        lista = []
        for id_prov, proveedor in self.proveedores.items():
            lista.append({
                'id_proveedor': id_prov,
                'alias': proveedor.get('alias', ''),
                'razon_social': proveedor.get('razon_social', ''),
                'telefono': proveedor.get('telefono', '')
            })
        return lista
    
    def agregar_proveedor(self,
                         id_proveedor: str,
                         alias: str,
                         rfc: str,
                         razon_social: str,
                         personal: str,
                         telefono: str,
                         codigo_postal: str,
                         estado: str,
                         ciudad: str,
                         municipio: str,
                         colonia: str,
                         direccion: str,
                         fax: str = "",
                         correo: str = "",
                         pagina_web: str = "",
                         tipo_pago: str = "mensual",  # semanal, quincenal, mensual
                         condiciones: str = "") -> Tuple[bool, str]:
        """Agrega un nuevo proveedor con todos los campos"""
        id_upper = id_proveedor.upper()
        
        # Validaciones
        if id_upper in self.proveedores:
            return (False, "El ID de proveedor ya existe")
        
        if not alias.strip():
            return (False, "El alias es requerido")
        
        if not razon_social.strip():
            return (False, "La razón social es requerida")
        
        if rfc and not validar_rfc(rfc):
            return (False, "RFC inválido (dejar vacío si no aplica)")
        
        if correo and not validar_email(correo):
            return (False, "Correo electrónico inválido")
        
        if not telefono.strip():
            return (False, "El teléfono es requerido")
        
        # Validar tipo de pago
        tipos_validos = ["semanal", "quincenal", "mensual"]
        if tipo_pago.lower() not in tipos_validos:
            return (False, f"Tipo de pago inválido. Debe ser: {', '.join(tipos_validos)}")
        
        # Registrar proveedor
        self.proveedores[id_upper] = {
            'alias': alias,
            'rfc': rfc.upper() if rfc else "",
            'razon_social': razon_social,
            'personal': personal,
            'telefono': telefono,
            'codigo_postal': codigo_postal,
            'estado': estado,
            'ciudad': ciudad,
            'municipio': municipio,
            'colonia': colonia,
            'direccion': direccion,
            'fax': fax,
            'correo': correo,
            'pagina_web': pagina_web,
            'tipo_pago': tipo_pago.lower(),
            'condiciones': condiciones,
            'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'activo': True
        }
        
        self.guardar_proveedores()
        return (True, f"Proveedor '{alias}' agregado exitosamente")
    
    def editar_proveedor(self,
                        id_proveedor: str,
                        alias: str,
                        rfc: str,
                        razon_social: str,
                        personal: str,
                        telefono: str,
                        codigo_postal: str,
                        estado: str,
                        ciudad: str,
                        municipio: str,
                        colonia: str,
                        direccion: str,
                        fax: str = "",
                        correo: str = "",
                        pagina_web: str = "",
                        tipo_pago: str = "mensual",
                        condiciones: str = "",
                        activo: bool = True) -> Tuple[bool, str]:
        """Edita un proveedor existente"""
        id_upper = id_proveedor.upper()
        
        if id_upper not in self.proveedores:
            return (False, "Proveedor no encontrado")
        
        if not alias.strip():
            return (False, "El alias es requerido")
        
        if not razon_social.strip():
            return (False, "La razón social es requerida")
        
        if rfc and not validar_rfc(rfc):
            return (False, "RFC inválido (dejar vacío si no aplica)")
        
        if correo and not validar_email(correo):
            return (False, "Correo electrónico inválido")
        
        if not telefono.strip():
            return (False, "El teléfono es requerido")
        
        # Validar tipo de pago
        tipos_validos = ["semanal", "quincenal", "mensual"]
        if tipo_pago.lower() not in tipos_validos:
            return (False, f"Tipo de pago inválido. Debe ser: {', '.join(tipos_validos)}")
        
        # Obtener datos existentes
        proveedor_existente = self.proveedores[id_upper]
        
        # Actualizar proveedor
        self.proveedores[id_upper] = {
            'alias': alias,
            'rfc': rfc.upper() if rfc else "",
            'razon_social': razon_social,
            'personal': personal,
            'telefono': telefono,
            'codigo_postal': codigo_postal,
            'estado': estado,
            'ciudad': ciudad,
            'municipio': municipio,
            'colonia': colonia,
            'direccion': direccion,
            'fax': fax,
            'correo': correo,
            'pagina_web': pagina_web,
            'tipo_pago': tipo_pago.lower(),
            'condiciones': condiciones,
            'fecha_registro': proveedor_existente.get('fecha_registro', 
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'activo': activo
        }
        
        self.guardar_proveedores()
        return (True, f"Proveedor '{alias}' actualizado exitosamente")
    
    def eliminar_proveedor(self, id_proveedor: str) -> Tuple[bool, str]:
        """Elimina (desactiva) un proveedor"""
        id_upper = id_proveedor.upper()
        
        if id_upper in self.proveedores:
            alias = self.proveedores[id_upper].get('alias', '')
            # En lugar de eliminar, marcamos como inactivo
            self.proveedores[id_upper]['activo'] = False
            self.proveedores[id_upper]['fecha_actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.guardar_proveedores()
            return (True, f"Proveedor '{alias}' marcado como inactivo")
        return (False, "Proveedor no encontrado")
    
    def eliminar_permanentemente(self, id_proveedor: str) -> Tuple[bool, str]:
        """Elimina permanentemente un proveedor"""
        id_upper = id_proveedor.upper()
        
        if id_upper in self.proveedores:
            alias = self.proveedores[id_upper].get('alias', '')
            del self.proveedores[id_upper]
            self.guardar_proveedores()
            return (True, f"Proveedor '{alias}' eliminado permanentemente")
        return (False, "Proveedor no encontrado")
    
    def buscar_proveedor(self, criterio: str, solo_activos: bool = True) -> List[Dict]:
        """Busca proveedores por múltiples criterios"""
        criterio_lower = criterio.lower()
        resultados = []
        
        for id_prov, proveedor in self.proveedores.items():
            # Filtrar por activos si se solicita
            if solo_activos and not proveedor.get('activo', True):
                continue
            
            if (criterio_lower in id_prov.lower() or
                criterio_lower in proveedor.get('alias', '').lower() or
                criterio_lower in proveedor.get('razon_social', '').lower() or
                criterio_lower in proveedor.get('rfc', '').lower() or
                criterio_lower in proveedor.get('personal', '').lower() or
                criterio_lower in proveedor.get('telefono', '') or
                criterio_lower in proveedor.get('correo', '').lower()):
                
                resultados.append({
                    'id_proveedor': id_prov,
                    **proveedor
                })
        
        return resultados
    
    def obtener_proveedores_activos(self) -> List[Dict]:
        """Obtiene solo proveedores activos"""
        return [
            {'id_proveedor': id_prov, **prov}
            for id_prov, prov in self.proveedores.items()
            if prov.get('activo', True)
        ]
    
    def exportar_proveedores_csv(self, ruta: str = 'proveedores_exportados.csv', solo_activos: bool = True) -> Tuple[bool, str]:
        """Exporta proveedores a un archivo CSV"""
        try:
            proveedores_a_exportar = self.proveedores
            
            if solo_activos:
                proveedores_a_exportar = {
                    id_prov: prov 
                    for id_prov, prov in self.proveedores.items() 
                    if prov.get('activo', True)
                }
            
            if not proveedores_a_exportar:
                return (False, "No hay proveedores para exportar")
            
            # Definir campos en orden específico
            campos = [
                'ID Proveedor',
                'Alias',
                'RFC',
                'Razon Social',
                'Personal',
                'Telefono',
                'Codigo Postal',
                'Estado',
                'Ciudad',
                'Municipio',
                'Colonia',
                'Direccion',
                'Fax',
                'Correo',
                'Pagina Web',
                'Tipo Pago',
                'Condiciones',
                'Activo',
                'Fecha Registro',
                'Fecha Actualizacion'
            ]
            
            with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=campos)
                writer.writeheader()
                
                for id_prov, proveedor in proveedores_a_exportar.items():
                    writer.writerow({
                        'ID Proveedor': id_prov,
                        'Alias': proveedor.get('alias', ''),
                        'RFC': proveedor.get('rfc', ''),
                        'Razon Social': proveedor.get('razon_social', ''),
                        'Personal': proveedor.get('personal', ''),
                        'Telefono': proveedor.get('telefono', ''),
                        'Codigo Postal': proveedor.get('codigo_postal', ''),
                        'Estado': proveedor.get('estado', ''),
                        'Ciudad': proveedor.get('ciudad', ''),
                        'Municipio': proveedor.get('municipio', ''),
                        'Colonia': proveedor.get('colonia', ''),
                        'Direccion': proveedor.get('direccion', ''),
                        'Fax': proveedor.get('fax', ''),
                        'Correo': proveedor.get('correo', ''),
                        'Pagina Web': proveedor.get('pagina_web', ''),
                        'Tipo Pago': proveedor.get('tipo_pago', ''),
                        'Condiciones': proveedor.get('condiciones', ''),
                        'Activo': 'SI' if proveedor.get('activo', True) else 'NO',
                        'Fecha Registro': proveedor.get('fecha_registro', ''),
                        'Fecha Actualizacion': proveedor.get('fecha_actualizacion', '')
                    })
            
            return (True, f"Proveedores exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar proveedores: {str(e)}")
    
    def exportar_proveedores_json(self, ruta: str = 'proveedores_exportados.json', solo_activos: bool = True) -> Tuple[bool, str]:
        """Exporta proveedores a un archivo JSON"""
        try:
            proveedores_a_exportar = self.proveedores
            
            if solo_activos:
                proveedores_a_exportar = {
                    id_prov: prov 
                    for id_prov, prov in self.proveedores.items() 
                    if prov.get('activo', True)
                }
            
            if not proveedores_a_exportar:
                return (False, "No hay proveedores para exportar")
            
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(proveedores_a_exportar, f, ensure_ascii=False, indent=2)
            
            return (True, f"Proveedores exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar proveedores: {str(e)}")
    
    def importar_proveedores_json(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa proveedores desde un archivo JSON"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            with open(ruta, 'r', encoding='utf-8') as f:
                datos_importados = json.load(f)
            
            if not isinstance(datos_importados, dict):
                return (False, "Formato de archivo inválido. Debe ser un objeto JSON con IDs como claves.")
            
            # Validar estructura
            proveedores_validos = 0
            proveedores_invalidos = 0
            
            for id_prov, proveedor in datos_importados.items():
                id_upper = id_prov.upper()
                
                # Validar campos requeridos
                campos_requeridos = ['alias', 'razon_social', 'telefono']
                if all(campo in proveedor for campo in campos_requeridos):
                    proveedores_validos += 1
                    
                    if sobrescribir or id_upper not in self.proveedores:
                        self.proveedores[id_upper] = proveedor
                        # Actualizar fechas
                        self.proveedores[id_upper]['fecha_actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        if 'fecha_registro' not in self.proveedores[id_upper]:
                            self.proveedores[id_upper]['fecha_registro'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        # Asegurar campo activo
                        if 'activo' not in self.proveedores[id_upper]:
                            self.proveedores[id_upper]['activo'] = True
                else:
                    proveedores_invalidos += 1
            
            self.guardar_proveedores()
            
            mensaje = f"Importación completada: {proveedores_validos} proveedores válidos importados"
            if proveedores_invalidos > 0:
                mensaje += f", {proveedores_invalidos} proveedores omitidos (campos faltantes)"
            
            return (True, mensaje)
        
        except json.JSONDecodeError as e:
            return (False, f"Error en formato JSON: {str(e)}")
        except Exception as e:
            return (False, f"Error al importar proveedores: {str(e)}")
    
    def importar_proveedores_csv(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa proveedores desde un archivo CSV"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            proveedores_importados = 0
            proveedores_omitidos = 0
            errores = []
            
            with open(ruta, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                # Verificar campos mínimos - solo ID Proveedor es obligatorio
                campos_requeridos = ['ID Proveedor']
                if not all(campo in reader.fieldnames for campo in campos_requeridos):
                    return (False, f"El CSV debe contener los campos: {', '.join(campos_requeridos)}")
                
                for i, fila in enumerate(reader, start=2):
                    try:
                        id_prov = fila.get('ID Proveedor', '').strip().upper()
                        
                        # Solo validar que ID de proveedor no sea vacío
                        if not id_prov:
                            errores.append(f"Línea {i}: ID de proveedor vacío")
                            proveedores_omitidos += 1
                            continue
                        
                        # Obtener todos los campos, permitiendo blancos
                        alias = fila.get('Alias', '').strip()
                        razon_social = fila.get('Razon Social', '').strip()
                        telefono = fila.get('Telefono', '').strip()
                        
                        # Solo importar si no existe o si se permite sobrescribir
                        if sobrescribir or id_prov not in self.proveedores:
                            self.proveedores[id_prov] = {
                                'alias': alias,
                                'rfc': fila.get('RFC', '').strip().upper(),
                                'razon_social': razon_social,
                                'personal': fila.get('Personal', '').strip(),
                                'telefono': telefono,
                                'codigo_postal': fila.get('Codigo Postal', '').strip(),
                                'estado': fila.get('Estado', '').strip(),
                                'ciudad': fila.get('Ciudad', '').strip(),
                                'municipio': fila.get('Municipio', '').strip(),
                                'colonia': fila.get('Colonia', '').strip(),
                                'direccion': fila.get('Direccion', '').strip(),
                                'fax': fila.get('Fax', '').strip(),
                                'correo': fila.get('Correo', '').strip(),
                                'pagina_web': fila.get('Pagina Web', '').strip(),
                                'tipo_pago': fila.get('Tipo Pago', 'mensual').strip().lower(),
                                'condiciones': fila.get('Condiciones', '').strip(),
                                'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'activo': True if fila.get('Activo', 'SI').strip().upper() == 'SI' else False
                            }
                            proveedores_importados += 1
                        else:
                            proveedores_omitidos += 1
                            
                    except Exception as e:
                        errores.append(f"Línea {i}: Error procesando fila - {str(e)}")
                        proveedores_omitidos += 1
            
            if proveedores_importados > 0:
                self.guardar_proveedores()
            
            mensaje = f"Importación CSV completada: {proveedores_importados} proveedores importados, {proveedores_omitidos} omitidos"
            if errores:
                mensaje += f"\nErrores encontrados:\n" + "\n".join(errores[:10])
            
            return (True if proveedores_importados > 0 else False, mensaje)
        
        except Exception as e:
            return (False, f"Error al importar CSV: {str(e)}")
    
    def generar_plantilla_csv(self, ruta: str = 'plantilla_proveedores.csv') -> Tuple[bool, str]:
        """Genera un archivo CSV de plantilla con los campos requeridos"""
        try:
            campos = [
                'ID Proveedor',
                'Alias',
                'RFC',
                'Razon Social',
                'Personal',
                'Telefono',
                'Codigo Postal',
                'Estado',
                'Ciudad',
                'Municipio',
                'Colonia',
                'Direccion',
                'Fax',
                'Correo',
                'Pagina Web',
                'Tipo Pago',
                'Condiciones',
                'Activo'
            ]
            
            with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(campos)
                # Agregar una fila de ejemplo
                writer.writerow([
                    'PROV001',
                    'Distribuidora Central',
                    'DCM800101TLA',
                    'Distribuidora Comercial Mexicana SA de CV',
                    'Juan Pérez',
                    '5551234567',
                    '01000',
                    'Ciudad de México',
                    'Ciudad de México',
                    'Cuauhtémoc',
                    'Centro',
                    'Av. Principal 123',
                    '5551234568',
                    'contacto@distribuidora.com',
                    'www.distribuidora.com',
                    'mensual',
                    'Pago a 30 días',
                    'SI'
                ])
            
            return (True, f"Plantilla generada: {ruta}")
        
        except Exception as e:
            return (False, f"Error al generar plantilla: {str(e)}")
    
    def exportar_proveedores_xlsx(self, ruta: str = 'proveedores_exportados.xlsx', solo_activos: bool = True) -> Tuple[bool, str]:
        """Exporta proveedores a un archivo XLSX"""
        try:
            proveedores_a_exportar = self.proveedores
            
            if solo_activos:
                proveedores_a_exportar = {
                    id_prov: prov 
                    for id_prov, prov in self.proveedores.items() 
                    if prov.get('activo', True)
                }
            
            if not proveedores_a_exportar:
                return (False, "No hay proveedores para exportar")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Proveedores"
            
            # Encabezados
            campos = ['ID Proveedor', 'Alias', 'RFC', 'Razon Social', 'Personal', 'Telefono',
                     'Codigo Postal', 'Estado', 'Ciudad', 'Municipio', 'Colonia', 'Direccion',
                     'Fax', 'Correo', 'Pagina Web', 'Tipo Pago', 'Condiciones', 'Activo']
            ws.append(campos)
            
            # Datos
            for id_prov, proveedor in proveedores_a_exportar.items():
                ws.append([
                    id_prov,
                    proveedor.get('alias', ''),
                    proveedor.get('rfc', ''),
                    proveedor.get('razon_social', ''),
                    proveedor.get('personal', ''),
                    proveedor.get('telefono', ''),
                    proveedor.get('codigo_postal', ''),
                    proveedor.get('estado', ''),
                    proveedor.get('ciudad', ''),
                    proveedor.get('municipio', ''),
                    proveedor.get('colonia', ''),
                    proveedor.get('direccion', ''),
                    proveedor.get('fax', ''),
                    proveedor.get('correo', ''),
                    proveedor.get('pagina_web', ''),
                    proveedor.get('tipo_pago', ''),
                    proveedor.get('condiciones', ''),
                    'SI' if proveedor.get('activo', True) else 'NO'
                ])
            
            wb.save(ruta)
            return (True, f"Proveedores exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar proveedores a XLSX: {str(e)}")
    
    def importar_proveedores_xlsx(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa proveedores desde un archivo XLSX"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            proveedores_importados = 0
            proveedores_omitidos = 0
            errores = []
            
            wb = load_workbook(ruta)
            ws = wb.active
            
            # Obtener encabezados
            headers = [cell.value for cell in ws[1]]
            
            # Validar campos mínimos
            campos_requeridos = ['ID Proveedor']
            if not all(campo in headers for campo in campos_requeridos):
                return (False, f"El XLSX debe contener los campos: {', '.join(campos_requeridos)}")
            
            # Procesar filas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear diccionario de fila
                    fila = dict(zip(headers, row))
                    
                    id_prov = str(fila.get('ID Proveedor', '')).strip().upper() if fila.get('ID Proveedor') else ''
                    
                    # Validación
                    if not id_prov:
                        errores.append(f"Línea {i}: ID de proveedor vacío")
                        proveedores_omitidos += 1
                        continue
                    
                    # Obtener campos opcionales
                    alias = str(fila.get('Alias', '')).strip() if fila.get('Alias') else ''
                    razon_social = str(fila.get('Razon Social', '')).strip() if fila.get('Razon Social') else ''
                    telefono = str(fila.get('Telefono', '')).strip() if fila.get('Telefono') else ''
                    
                    # Importar proveedor
                    if sobrescribir or id_prov not in self.proveedores:
                        self.proveedores[id_prov] = {
                            'alias': alias,
                            'rfc': str(fila.get('RFC', '')).strip().upper() if fila.get('RFC') else '',
                            'razon_social': razon_social,
                            'personal': str(fila.get('Personal', '')).strip() if fila.get('Personal') else '',
                            'telefono': telefono,
                            'codigo_postal': str(fila.get('Codigo Postal', '')).strip() if fila.get('Codigo Postal') else '',
                            'estado': str(fila.get('Estado', '')).strip() if fila.get('Estado') else '',
                            'ciudad': str(fila.get('Ciudad', '')).strip() if fila.get('Ciudad') else '',
                            'municipio': str(fila.get('Municipio', '')).strip() if fila.get('Municipio') else '',
                            'colonia': str(fila.get('Colonia', '')).strip() if fila.get('Colonia') else '',
                            'direccion': str(fila.get('Direccion', '')).strip() if fila.get('Direccion') else '',
                            'fax': str(fila.get('Fax', '')).strip() if fila.get('Fax') else '',
                            'correo': str(fila.get('Correo', '')).strip() if fila.get('Correo') else '',
                            'pagina_web': str(fila.get('Pagina Web', '')).strip() if fila.get('Pagina Web') else '',
                            'tipo_pago': str(fila.get('Tipo Pago', 'mensual')).strip().lower() if fila.get('Tipo Pago') else 'mensual',
                            'condiciones': str(fila.get('Condiciones', '')).strip() if fila.get('Condiciones') else '',
                            'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'activo': True if str(fila.get('Activo', 'SI')).strip().upper() == 'SI' else False
                        }
                        proveedores_importados += 1
                    else:
                        proveedores_omitidos += 1
                        
                except Exception as e:
                    errores.append(f"Línea {i}: Error procesando fila - {str(e)}")
                    proveedores_omitidos += 1
            
            if proveedores_importados > 0:
                self.guardar_proveedores()
            
            mensaje = f"Importación XLSX completada: {proveedores_importados} proveedores importados, {proveedores_omitidos} omitidos"
            if errores:
                mensaje += f"\nErrores encontrados:\n" + "\n".join(errores[:10])
            
            return (True if proveedores_importados > 0 else False, mensaje)
        
        except Exception as e:
            return (False, f"Error al importar XLSX: {str(e)}")
    
    def obtener_proveedores_por_tipo_pago(self, tipo_pago: str) -> List[Dict]:
        """Obtiene proveedores por tipo de pago"""
        tipos_validos = ["semanal", "quincenal", "mensual"]
        if tipo_pago.lower() not in tipos_validos:
            return []
        
        return [
            {'id_proveedor': id_prov, **prov}
            for id_prov, prov in self.proveedores.items()
            if prov.get('tipo_pago', '').lower() == tipo_pago.lower() and prov.get('activo', True)
        ]
    
    def activar_proveedor(self, id_proveedor: str) -> Tuple[bool, str]:
        """Reactiva un proveedor inactivo"""
        id_upper = id_proveedor.upper()
        
        if id_upper in self.proveedores:
            alias = self.proveedores[id_upper].get('alias', '')
            self.proveedores[id_upper]['activo'] = True
            self.proveedores[id_upper]['fecha_actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.guardar_proveedores()
            return (True, f"Proveedor '{alias}' reactivado exitosamente")
        return (False, "Proveedor no encontrado")
    
    # Método de compatibilidad
    def obtener_lista_nombres(self) -> List[str]:
        """Método de compatibilidad para obtener lista de nombres (legacy)"""
        return [f"{id_prov} - {prov.get('alias', '')}" for id_prov, prov in self.proveedores.items() if prov.get('activo', True)]


class GestorClientes:
    """Clase para gestionar los clientes con todos los campos fiscales"""
    def __init__(self):
        self.clientes = {}
        self.cargar_clientes()
    
    def cargar_clientes(self, archivo: str = 'clientes.json') -> None:
        """Carga los clientes desde archivo JSON"""
        if os.path.exists(archivo):
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    self.clientes = json.load(f)
            except Exception as e:
                print(f"Error al cargar clientes: {e}")
    
    def guardar_clientes(self, archivo: str = 'clientes.json') -> None:
        """Guarda los clientes en archivo JSON"""
        try:
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump(self.clientes, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error al guardar clientes: {e}")
    
    def obtener_cliente(self, rfc: str) -> Optional[Dict]:
        """Obtiene un cliente por RFC"""
        return self.clientes.get(rfc.upper())
    
    def obtener_todos(self) -> Dict:
        """Obtiene todos los clientes"""
        return self.clientes
    
    def obtener_lista_simplificada(self) -> List[Dict]:
        """Obtiene lista de clientes para combos/selects"""
        lista = []
        for rfc, cliente in self.clientes.items():
            lista.append({
                'rfc': rfc,
                'razon_social': cliente.get('razon_social', ''),
                'telefono': cliente.get('telefono', ''),
                'correo': cliente.get('correo', '')
            })
        return lista
    
    def agregar_cliente(self,
                       rfc: str,
                       razon_social: str,
                       regimen_fiscal: str,
                       uso_cfdi: str,
                       codigo_postal: str,
                       direccion_fiscal: str,
                       estado: str,
                       ciudad: str,
                       municipio: str,
                       colonia: str,
                       telefono: str,
                       correo: str) -> Tuple[bool, str]:
        """Agrega un nuevo cliente con todos los campos fiscales"""
        rfc_upper = rfc.upper()
        
        # Validaciones
        if rfc_upper in self.clientes:
            return (False, "El cliente con este RFC ya existe")
        
        if not validar_rfc(rfc_upper):
            return (False, "RFC inválido")
        
        if not validar_email(correo):
            return (False, "Correo electrónico inválido")
        
        if not telefono.strip():
            return (False, "El teléfono es requerido")
        
        if not codigo_postal.strip():
            return (False, "El código postal es requerido")
        
        # Registrar cliente
        self.clientes[rfc_upper] = {
            'razon_social': razon_social,
            'regimen_fiscal': regimen_fiscal,
            'uso_cfdi': uso_cfdi,
            'codigo_postal': codigo_postal,
            'direccion_fiscal': direccion_fiscal,
            'estado': estado,
            'ciudad': ciudad,
            'municipio': municipio,
            'colonia': colonia,
            'telefono': telefono,
            'correo': correo,
            'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        self.guardar_clientes()
        return (True, f"Cliente '{razon_social}' agregado exitosamente")
    
    def editar_cliente(self,
                      rfc: str,
                      razon_social: str,
                      regimen_fiscal: str,
                      uso_cfdi: str,
                      codigo_postal: str,
                      direccion_fiscal: str,
                      estado: str,
                      ciudad: str,
                      municipio: str,
                      colonia: str,
                      telefono: str,
                      correo: str) -> Tuple[bool, str]:
        """Edita un cliente existente"""
        rfc_upper = rfc.upper()
        
        if rfc_upper not in self.clientes:
            return (False, "Cliente no encontrado")
        
        if not validar_email(correo):
            return (False, "Correo electrónico inválido")
        
        if not telefono.strip():
            return (False, "El teléfono es requerido")
        
        if not codigo_postal.strip():
            return (False, "El código postal es requerido")
        
        # Obtener datos existentes
        cliente_existente = self.clientes[rfc_upper]
        
        # Actualizar cliente
        self.clientes[rfc_upper] = {
            'razon_social': razon_social,
            'regimen_fiscal': regimen_fiscal,
            'uso_cfdi': uso_cfdi,
            'codigo_postal': codigo_postal,
            'direccion_fiscal': direccion_fiscal,
            'estado': estado,
            'ciudad': ciudad,
            'municipio': municipio,
            'colonia': colonia,
            'telefono': telefono,
            'correo': correo,
            'fecha_registro': cliente_existente.get('fecha_registro', 
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        self.guardar_clientes()
        return (True, f"Cliente '{razon_social}' actualizado exitosamente")
    
    def eliminar_cliente(self, rfc: str) -> Tuple[bool, str]:
        """Elimina un cliente"""
        rfc_upper = rfc.upper()
        
        if rfc_upper in self.clientes:
            razon_social = self.clientes[rfc_upper].get('razon_social', '')
            del self.clientes[rfc_upper]
            self.guardar_clientes()
            return (True, f"Cliente '{razon_social}' eliminado exitosamente")
        return (False, "Cliente no encontrado")
    
    def buscar_cliente(self, criterio: str) -> List[Dict]:
        """Busca clientes por RFC, razón social o correo"""
        criterio_lower = criterio.lower()
        resultados = []
        
        for rfc, cliente in self.clientes.items():
            if (criterio_lower in rfc.lower() or
                criterio_lower in cliente.get('razon_social', '').lower() or
                criterio_lower in cliente.get('correo', '').lower() or
                criterio_lower in cliente.get('telefono', '')):
                
                resultados.append({
                    'rfc': rfc,
                    **cliente
                })
        
        return resultados
    
    def exportar_clientes_csv(self, ruta: str = 'clientes_exportados.csv') -> Tuple[bool, str]:
        """Exporta todos los clientes a un archivo CSV"""
        try:
            if not self.clientes:
                return (False, "No hay clientes para exportar")
            
            # Definir campos en orden específico
            campos = [
                'RFC',
                'Razon Social',
                'Regimen Fiscal',
                'Uso CFDI',
                'Codigo Postal',
                'Direccion Fiscal',
                'Estado',
                'Ciudad',
                'Municipio',
                'Colonia',
                'Telefono',
                'Correo',
                'Fecha Registro',
                'Fecha Actualizacion'
            ]
            
            with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=campos)
                writer.writeheader()
                
                for rfc, cliente in self.clientes.items():
                    writer.writerow({
                        'RFC': rfc,
                        'Razon Social': cliente.get('razon_social', ''),
                        'Regimen Fiscal': cliente.get('regimen_fiscal', ''),
                        'Uso CFDI': cliente.get('uso_cfdi', ''),
                        'Codigo Postal': cliente.get('codigo_postal', ''),
                        'Direccion Fiscal': cliente.get('direccion_fiscal', ''),
                        'Estado': cliente.get('estado', ''),
                        'Ciudad': cliente.get('ciudad', ''),
                        'Municipio': cliente.get('municipio', ''),
                        'Colonia': cliente.get('colonia', ''),
                        'Telefono': cliente.get('telefono', ''),
                        'Correo': cliente.get('correo', ''),
                        'Fecha Registro': cliente.get('fecha_registro', ''),
                        'Fecha Actualizacion': cliente.get('fecha_actualizacion', '')
                    })
            
            return (True, f"Clientes exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar clientes: {str(e)}")
    
    def exportar_clientes_json(self, ruta: str = 'clientes_exportados.json') -> Tuple[bool, str]:
        """Exporta todos los clientes a un archivo JSON"""
        try:
            if not self.clientes:
                return (False, "No hay clientes para exportar")
            
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(self.clientes, f, ensure_ascii=False, indent=2)
            
            return (True, f"Clientes exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar clientes: {str(e)}")
    
    def importar_clientes_json(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa clientes desde un archivo JSON"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            with open(ruta, 'r', encoding='utf-8') as f:
                datos_importados = json.load(f)
            
            if not isinstance(datos_importados, dict):
                return (False, "Formato de archivo inválido. Debe ser un objeto JSON con RFCs como claves.")
            
            # Validar estructura de cada cliente
            clientes_validos = 0
            clientes_invalidos = 0
            
            for rfc, cliente in datos_importados.items():
                if not validar_rfc(rfc):
                    clientes_invalidos += 1
                    continue
                
                # Validar campos requeridos
                campos_requeridos = ['razon_social', 'telefono', 'codigo_postal']
                if all(campo in cliente for campo in campos_requeridos):
                    clientes_validos += 1
                    
                    if sobrescribir or rfc.upper() not in self.clientes:
                        self.clientes[rfc.upper()] = cliente
                        # Actualizar fechas
                        self.clientes[rfc.upper()]['fecha_actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        if 'fecha_registro' not in self.clientes[rfc.upper()]:
                            self.clientes[rfc.upper()]['fecha_registro'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    clientes_invalidos += 1
            
            self.guardar_clientes()
            
            mensaje = f"Importación completada: {clientes_validos} clientes válidos importados"
            if clientes_invalidos > 0:
                mensaje += f", {clientes_invalidos} clientes omitidos (RFC inválido o campos faltantes)"
            
            return (True, mensaje)
        
        except json.JSONDecodeError as e:
            return (False, f"Error en formato JSON: {str(e)}")
        except Exception as e:
            return (False, f"Error al importar clientes: {str(e)}")
    
    def importar_clientes_csv(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa clientes desde un archivo CSV"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            clientes_importados = 0
            clientes_omitidos = 0
            errores = []
            
            with open(ruta, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                # Verificar campos mínimos - solo RFC y Razon Social son obligatorios
                campos_requeridos = ['RFC', 'Razon Social']
                if not all(campo in reader.fieldnames for campo in campos_requeridos):
                    return (False, f"El CSV debe contener los campos: {', '.join(campos_requeridos)}")
                
                for i, fila in enumerate(reader, start=2):  # start=2 porque la fila 1 es el encabezado
                    try:
                        rfc = fila.get('RFC', '').strip().upper()
                        razon_social = fila.get('Razon Social', '').strip()
                        codigo_postal = fila.get('Codigo Postal', '').strip()
                        
                        # Validaciones básicas - RFC puede ser genérico o XAXX010101000
                        if not rfc:
                            errores.append(f"Línea {i}: RFC vacío")
                            clientes_omitidos += 1
                            continue
                        
                        # Aceptar RFCs genéricos también
                        if not validar_rfc(rfc) and rfc.upper() != 'XAXX010101000':
                            errores.append(f"Línea {i}: RFC inválido")
                            clientes_omitidos += 1
                            continue
                        
                        if not razon_social:
                            errores.append(f"Línea {i}: Razón social vacía")
                            clientes_omitidos += 1
                            continue
                        
                        # Solo importar si no existe o si se permite sobrescribir
                        if sobrescribir or rfc not in self.clientes:
                            self.clientes[rfc] = {
                                'razon_social': razon_social,
                                'regimen_fiscal': fila.get('Regimen Fiscal', '').strip(),
                                'uso_cfdi': fila.get('Uso CFDI', '').strip(),
                                'codigo_postal': codigo_postal,
                                'direccion_fiscal': fila.get('Direccion Fiscal', '').strip(),
                                'estado': fila.get('Estado', '').strip(),
                                'ciudad': fila.get('Ciudad', '').strip(),
                                'municipio': fila.get('Municipio', '').strip(),
                                'colonia': fila.get('Colonia', '').strip(),
                                'telefono': fila.get('Telefono', '').strip(),
                                'correo': fila.get('Correo', '').strip(),
                                'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            clientes_importados += 1
                        else:
                            clientes_omitidos += 1
                            
                    except Exception as e:
                        errores.append(f"Línea {i}: Error procesando fila - {str(e)}")
                        clientes_omitidos += 1
            
            if clientes_importados > 0:
                self.guardar_clientes()
            
            mensaje = f"Importación CSV completada: {clientes_importados} clientes importados, {clientes_omitidos} omitidos"
            if errores:
                mensaje += f"\nErrores encontrados:\n" + "\n".join(errores[:10])  # Mostrar solo primeros 10 errores
            
            return (True if clientes_importados > 0 else False, mensaje)
        
        except Exception as e:
            return (False, f"Error al importar CSV: {str(e)}")
    
    def generar_plantilla_csv(self, ruta: str = 'plantilla_clientes.csv') -> Tuple[bool, str]:
        """Genera un archivo CSV de plantilla con los campos requeridos"""
        try:
            campos = [
                'RFC',
                'Razon Social',
                'Regimen Fiscal',
                'Uso CFDI',
                'Codigo Postal',
                'Direccion Fiscal',
                'Estado',
                'Ciudad',
                'Municipio',
                'Colonia',
                'Telefono',
                'Correo'
            ]
            
            with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(campos)
                # Agregar una fila de ejemplo
                writer.writerow([
                    'XAXX010101000',
                    'EJEMPLO SA DE CV',
                    '601 - General de Ley Personas Morales',
                    'G03 - Gastos en general',
                    '01000',
                    'Calle Ejemplo 123',
                    'Ciudad de México',
                    'Ciudad de México',
                    'Cuauhtémoc',
                    'Centro',
                    '5551234567',
                    'ejemplo@empresa.com'
                ])
            
            return (True, f"Plantilla generada: {ruta}")
        
        except Exception as e:
            return (False, f"Error al generar plantilla: {str(e)}")
    
    def exportar_clientes_xlsx(self, ruta: str = 'clientes_exportados.xlsx') -> Tuple[bool, str]:
        """Exporta clientes a un archivo XLSX"""
        try:
            if not self.clientes:
                return (False, "No hay clientes para exportar")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
            
            # Encabezados
            campos = ['RFC', 'Razon Social', 'Regimen Fiscal', 'Uso CFDI', 'Codigo Postal',
                     'Direccion Fiscal', 'Estado', 'Ciudad', 'Municipio', 'Colonia', 'Telefono', 'Correo']
            ws.append(campos)
            
            # Datos
            for rfc, cliente in self.clientes.items():
                ws.append([
                    rfc,
                    cliente.get('razon_social', ''),
                    cliente.get('regimen_fiscal', ''),
                    cliente.get('uso_cfdi', ''),
                    cliente.get('codigo_postal', ''),
                    cliente.get('direccion_fiscal', ''),
                    cliente.get('estado', ''),
                    cliente.get('ciudad', ''),
                    cliente.get('municipio', ''),
                    cliente.get('colonia', ''),
                    cliente.get('telefono', ''),
                    cliente.get('correo', '')
                ])
            
            wb.save(ruta)
            return (True, f"Clientes exportados exitosamente: {ruta}")
        
        except Exception as e:
            return (False, f"Error al exportar clientes a XLSX: {str(e)}")
    
    def importar_clientes_xlsx(self, ruta: str, sobrescribir: bool = False) -> Tuple[bool, str]:
        """Importa clientes desde un archivo XLSX"""
        try:
            if not os.path.exists(ruta):
                return (False, f"Archivo no encontrado: {ruta}")
            
            clientes_importados = 0
            clientes_omitidos = 0
            errores = []
            
            wb = load_workbook(ruta)
            ws = wb.active
            
            # Obtener encabezados
            headers = [cell.value for cell in ws[1]]
            
            # Validar campos mínimos
            campos_requeridos = ['RFC', 'Razon Social']
            if not all(campo in headers for campo in campos_requeridos):
                return (False, f"El XLSX debe contener los campos: {', '.join(campos_requeridos)}")
            
            # Procesar filas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear diccionario de fila
                    fila = dict(zip(headers, row))
                    
                    rfc = str(fila.get('RFC', '')).strip().upper() if fila.get('RFC') else ''
                    razon_social = str(fila.get('Razon Social', '')).strip() if fila.get('Razon Social') else ''
                    codigo_postal = str(fila.get('Codigo Postal', '')).strip() if fila.get('Codigo Postal') else ''
                    
                    # Validaciones
                    if not rfc:
                        errores.append(f"Línea {i}: RFC vacío")
                        clientes_omitidos += 1
                        continue
                    
                    if not validar_rfc(rfc) and rfc.upper() != 'XAXX010101000':
                        errores.append(f"Línea {i}: RFC inválido")
                        clientes_omitidos += 1
                        continue
                    
                    if not razon_social:
                        errores.append(f"Línea {i}: Razón social vacía")
                        clientes_omitidos += 1
                        continue
                    
                    # Importar cliente
                    if sobrescribir or rfc not in self.clientes:
                        self.clientes[rfc] = {
                            'razon_social': razon_social,
                            'regimen_fiscal': str(fila.get('Regimen Fiscal', '')).strip() if fila.get('Regimen Fiscal') else '',
                            'uso_cfdi': str(fila.get('Uso CFDI', '')).strip() if fila.get('Uso CFDI') else '',
                            'codigo_postal': codigo_postal,
                            'direccion_fiscal': str(fila.get('Direccion Fiscal', '')).strip() if fila.get('Direccion Fiscal') else '',
                            'estado': str(fila.get('Estado', '')).strip() if fila.get('Estado') else '',
                            'ciudad': str(fila.get('Ciudad', '')).strip() if fila.get('Ciudad') else '',
                            'municipio': str(fila.get('Municipio', '')).strip() if fila.get('Municipio') else '',
                            'colonia': str(fila.get('Colonia', '')).strip() if fila.get('Colonia') else '',
                            'telefono': str(fila.get('Telefono', '')).strip() if fila.get('Telefono') else '',
                            'correo': str(fila.get('Correo', '')).strip() if fila.get('Correo') else '',
                            'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        clientes_importados += 1
                    else:
                        clientes_omitidos += 1
                        
                except Exception as e:
                    errores.append(f"Línea {i}: Error procesando fila - {str(e)}")
                    clientes_omitidos += 1
            
            if clientes_importados > 0:
                self.guardar_clientes()
            
            mensaje = f"Importación XLSX completada: {clientes_importados} clientes importados, {clientes_omitidos} omitidos"
            if errores:
                mensaje += f"\nErrores encontrados:\n" + "\n".join(errores[:10])
            
            return (True if clientes_importados > 0 else False, mensaje)
        
        except Exception as e:
            return (False, f"Error al importar XLSX: {str(e)}")


class GeneradorReportes:
    """Clase para generar reportes de ventas"""
    def __init__(self, gestor_ventas: GestorVentas):
        self.gestor_ventas = gestor_ventas
    
    def generar_reporte_diario(self, fecha: str = None) -> Dict:
        """Genera un reporte de ventas del dia"""
        if fecha is None:
            fecha = datetime.now().strftime('%Y-%m-%d')
        
        ventas = self.gestor_ventas.obtener_ventas()
        ventas_del_dia = [v for v in ventas if v.get('fecha', '').startswith(fecha)]
        
        total = sum(v.get('total', 0) for v in ventas_del_dia)
        return {
            'fecha': fecha,
            'cantidad_ventas': len(ventas_del_dia),
            'total': total
        }
    
    def generar_reporte_csv(self, archivo: str = 'reporte_ventas.csv') -> None:
        """Genera un reporte en formato CSV"""
        try:
            ventas = self.gestor_ventas.obtener_ventas()
            if not ventas:
                return
            
            # Crear una lista plana de productos
            rows = []
            for venta in ventas:
                for producto in venta.get('productos', []):
                    rows.append({
                        'folio': venta.get('folio', ''),
                        'fecha': venta.get('fecha', ''),
                        'producto': producto.get('nombre', ''),
                        'descripcion': producto.get('descripcion', ''),
                        'cantidad': producto.get('cantidad', 0),
                        'precio': producto.get('precio_aplicado', 0),
                        'subtotal': producto.get('subtotal', 0),
                        'es_mayoreo': 'SI' if producto.get('es_mayoreo', False) else 'NO'
                    })
            
            if rows:
                with open(archivo, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows(rows)
        except Exception as e:
            print(f"Error al generar reporte CSV: {e}")


class GeneradorTicket:
    """Clase para generar tickets de venta"""
    def __init__(self):
        pass
    
    @staticmethod
    def generar_ticket(venta: Dict) -> str:
        """Genera un ticket de venta con información de precios"""
        ticket = "=" * 50 + "\n"
        ticket += "TICKET DE VENTA\n"
        ticket += "=" * 50 + "\n"
        
        if 'fecha' in venta:
            ticket += f"Fecha: {venta['fecha']}\n"
        
        if 'folio' in venta:
            ticket += f"Folio: {venta['folio']}\n"
        
        ticket += "-" * 50 + "\n"
        ticket += "PRODUCTOS:\n"
        ticket += "-" * 50 + "\n"
        
        if 'productos' in venta:
            for prod in venta['productos']:
                precio_unitario = prod.get('precio_aplicado', 0)
                es_mayoreo = prod.get('es_mayoreo', False)
                precio_minorista = prod.get('precio_minorista', 0)
                
                ticket += f"{prod.get('nombre', 'Producto')}\n"
                if prod.get('descripcion'):
                    ticket += f"  Desc: {prod.get('descripcion')[:30]}...\n"
                ticket += f"  {prod.get('cantidad', 1)} x ${precio_unitario:.2f}"
                
                if es_mayoreo and precio_minorista > 0:
                    ahorro = (precio_minorista - precio_unitario) * prod.get('cantidad', 1)
                    ticket += f" (Mayoreo) Ahorro: ${ahorro:.2f}"
                
                subtotal = prod.get('subtotal', 0)
                ticket += f" = ${subtotal:.2f}\n"
                
                if prod.get('unidad', 'pz') != 'pz':
                    ticket += f"  Unidad: {prod.get('unidad')}\n"
        
        ticket += "-" * 50 + "\n"
        
        # Mostrar resumen de descuentos si aplica
        if 'detalle_precios' in venta and venta['detalle_precios']['descuento_total'] > 0:
            ticket += f"Descuento por mayoreo: ${venta['detalle_precios']['descuento_total']:.2f}\n"
        
        ticket += f"Subtotal: ${venta.get('subtotal', 0):.2f}\n"
        ticket += f"IVA (16%): ${venta.get('iva', 0):.2f}\n"
        ticket += f"TOTAL: ${venta.get('total', 0):.2f}\n"
        ticket += f"Pago: ${venta.get('pago', 0):.2f}\n"
        ticket += f"Cambio: ${venta.get('cambio', 0):.2f}\n"
        
        ticket += "=" * 50 + "\n"
        ticket += "¡GRACIAS POR SU COMPRA!\n"
        ticket += "=" * 50 + "\n"
        
        return ticket
    
    @staticmethod
    def guardar_ticket(venta: Dict, archivo: str = None) -> str:
        """Guarda un ticket en archivo y retorna la ruta"""
        if archivo is None:
            folio = venta.get('folio', datetime.now().strftime('%Y%m%d_%H%M%S'))
            archivo = f"ticket_{folio}.txt"
        
        try:
            with open(archivo, 'w', encoding='utf-8') as f:
                f.write(GeneradorTicket.generar_ticket(venta))
            return os.path.abspath(archivo)
        except Exception as e:
            print(f"Error al guardar ticket: {e}")
            return ""


# Funciones de ayuda para catalogos SAT
def obtener_regimenes_fiscales() -> List[Dict]:
    """Retorna lista de regímenes fiscales comunes del SAT"""
    return [
        {"clave": "601", "descripcion": "General de Ley Personas Morales"},
        {"clave": "603", "descripcion": "Personas Morales con Fines no Lucrativos"},
        {"clave": "605", "descripcion": "Sueldos y Salarios e Ingresos Asimilados a Salarios"},
        {"clave": "606", "descripcion": "Arrendamiento"},
        {"clave": "607", "descripcion": "Régimen de Enajenación o Adquisición de Bienes"},
        {"clave": "608", "descripcion": "Demás ingresos"},
        {"clave": "610", "descripcion": "Residentes en el Extranjero sin Establecimiento Permanente en México"},
        {"clave": "611", "descripcion": "Ingresos por Dividendos (socios y accionistas)"},
        {"clave": "612", "descripcion": "Personas Físicas con Actividades Empresariales y Profesionales"},
        {"clave": "614", "descripcion": "Ingresos por intereses"},
        {"clave": "615", "descripcion": "Régimen de los ingresos por obtención de premios"},
        {"clave": "616", "descripcion": "Sin obligaciones fiscales"},
        {"clave": "620", "descripcion": "Sociedades Cooperativas de Producción que optan por diferir sus ingresos"},
        {"clave": "621", "descripcion": "Incorporación Fiscal"},
        {"clave": "622", "descripcion": "Actividades Agrícolas, Ganaderas, Silvícolas y Pesqueras"},
        {"clave": "623", "descripcion": "Opcional para Grupos de Sociedades"},
        {"clave": "624", "descripcion": "Coordinados"},
        {"clave": "625", "descripcion": "Régimen de las Actividades Empresariales con ingresos a través de Plataformas Tecnológicas"},
        {"clave": "626", "descripcion": "Régimen Simplificado de Confianza"}
    ]

def obtener_usos_cfdi() -> List[Dict]:
    """Retorna lista de usos de CFDI comunes"""
    return [
        {"clave": "G01", "descripcion": "Adquisición de mercancías"},
        {"clave": "G02", "descripcion": "Devoluciones, descuentos o bonificaciones"},
        {"clave": "G03", "descripcion": "Gastos en general"},
        {"clave": "I01", "descripcion": "Construcciones"},
        {"clave": "I02", "descripcion": "Mobiliario y equipo de oficina por inversiones"},
        {"clave": "I03", "descripcion": "Equipo de transporte"},
        {"clave": "I04", "descripcion": "Equipo de cómputo y accesorios"},
        {"clave": "I05", "descripcion": "Dados, troqueles, moldes, matrices y herramental"},
        {"clave": "I06", "descripcion": "Comunicaciones telefónicas"},
        {"clave": "I07", "descripcion": "Comunicaciones satelitales"},
        {"clave": "I08", "descripcion": "Otra maquinaria y equipo"},
        {"clave": "D01", "descripcion": "Honorarios médicos, dentales y gastos hospitalarios"},
        {"clave": "D02", "descripcion": "Gastos médicos por incapacidad o discapacidad"},
        {"clave": "D03", "descripcion": "Gastos funerales"},
        {"clave": "D04", "descripcion": "Donativos"},
        {"clave": "D05", "descripcion": "Intereses reales efectivamente pagados por créditos hipotecarios"},
        {"clave": "D06", "descripcion": "Aportaciones voluntarias al SAR"},
        {"clave": "D07", "descripcion": "Primas por seguros de gastos médicos"},
        {"clave": "D08", "descripcion": "Gastos de transportación escolar obligatoria"},
        {"clave": "D09", "descripcion": "Depósitos en cuentas para el ahorro, primas que tengan como base planes de pensiones"},
        {"clave": "D10", "descripcion": "Pagos por servicios educativos (colegiaturas)"},
        {"clave": "S01", "descripcion": "Sin efectos fiscales"},
        {"clave": "CP01", "descripcion": "Pagos"},
        {"clave": "CN01", "descripcion": "Nómina"}
    ]


# Función de ejemplo para demostrar el uso
def ejemplo_punto_venta_descripcion():
    """Ejemplo de cómo usar el punto de venta con búsqueda por descripción"""
    
    # Crear gestores
    gestor_inventario = Gestor_Inventario()
    gestor_ventas = GestorVentas(gestor_inventario)
    
    # Ejemplo de agregar productos por descripción
    print("=== EJEMPLO PUNTO DE VENTA ===")
    
    # Buscar y agregar producto por descripción
    resultado = gestor_ventas.agregar_producto_por_descripcion("leche", 2)
    print(f"Resultado 1: {resultado[1]}")
    
    # Agregar más cantidad del mismo producto (ahora debería ser 5 en total)
    resultado = gestor_ventas.agregar_producto_por_descripcion("leche", 3)
    print(f"Resultado 2: {resultado[1]}")
    
    # Agregar otro producto (6 unidades para probar precio mayoreo)
    resultado = gestor_ventas.agregar_producto_por_descripcion("pan", 6)
    print(f"Resultado 3: {resultado[1]}")
    
    # Mostrar productos en la venta
    print("\nProductos en venta:")
    for prod in gestor_ventas.obtener_productos_venta():
        precio_unitario = prod.obtener_precio_unitario()
        print(f"- {prod.nombre}: {prod.cantidad} x ${precio_unitario:.2f} = ${prod.subtotal():.2f} "
              f"(Mayoreo: {'Sí' if prod.cantidad >= 6 else 'No'})")
    
    # Calcular totales
    print(f"\nSubtotal: ${gestor_ventas.calcular_subtotal():.2f}")
    print(f"IVA (16%): ${gestor_ventas.calcular_iva():.2f}")
    print(f"Total: ${gestor_ventas.calcular_total():.2f}")
    
    # Procesar venta
    pago = 500.00
    resultado_venta = gestor_ventas.procesar_venta(pago)
    
    if resultado_venta[0]:
        print(f"\nVenta procesada exitosamente!")
        print(f"Folio: {resultado_venta[2]['folio']}")
        print(f"Cambio: ${resultado_venta[2]['cambio']:.2f}")
        
        # Generar ticket
        ticket = GeneradorTicket.generar_ticket(resultado_venta[2])
        print("\n" + ticket)
    else:
        print(f"\nError: {resultado_venta[1]}")